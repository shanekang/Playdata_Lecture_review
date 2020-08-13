from openpyxl import load_workbook #openpyxl 모듈중에 load_workbook 을 사용한다. 
                                   #load_workbook은 엑셀 파일을 열 때 사용한다.
import json #json 파일을 열거나 만들 때 사용할 모듈
import datetime #datatime은 년, 월, 일을 통해 날짜를 표현하는 클래스 혹은 모듈이다.

with open("subway_location.json", "r", encoding="utf-8") as json_open :
    #json_open의 이름으로 subway_location.json을 읽어오는 코드이다.
    subway_location=json.load(json_open) #위에 with open구문으로 불러온 json파일의 데이터를
                                         #subway_location에 담아준다.

# print(subway_location['가락시장'][0])
# print(subway_location['가락시장'][1])

load_wb = load_workbook("subway_log_2019_1.xlsx", data_only=True)
#subway_log_2019.xlsx 엑셀 파일을 열어서 load_wb에 넣어준다. 
#여기서 data_only=True 해주어야만 나중에 .value로 엑셀의 수식이 아닌 셀 값만 받아오게 된다!
load_ws = load_wb['Sheet1']
#load_ws에 load_wb 셀 값중 Sheet1에 있는 데이터만 넣어준다.
#print(load_wb)
#print(load_ws)
#print(load_ws['B4'].value)
#print(load_ws['A3'].value) #.value를 넣어 주어야만 엑셀의 수식아닌 셀의 값을 마침내 받아오게 된다.

get_rows = load_ws.rows #.rows는 열을 한줄씩 가지고 온다

metro_2019_log = []

for row in get_rows : #for문으로 위에 한줄 씩 가지고 온 열들을 row에 차례로 넣어준다.
    rowinit = []
    for cell in row : #가지고 온 열들의 값을 하나씩 cell에 넣어준다.
       rowinit.append(cell.value) #for문으로 가지고 온 cell에 저장된 열 값중에 value 값만 
                                  #위에 rowinit 리스트에 차례로 넣어준다.
    metro_2019_log.append(rowinit) #한줄 씩 열의 value값이 들은 rowinit 리스트의 모든 값을 metro_2019_log에 통채로 넣어준다.


#print(metro_2019_log)
#get_cells = load_ws['A1':'Z1']
#print(get_cells[0][0].value)


#for row in get_cells :
#     for cell in row :
#         print(cell.value)

#metro_2019_log[0][3]


# i = 0
# indata = 0
# outdata = 1

# i = 2
# indata = 2  => 2
# outdata = 3 => 3

# i = 4

log_count=len(metro_2019_log)#위에 metro_2019_log에 넣은 모든 열들의 갯수를 log_count(550)에 넣어준다.

# 1호선 => Line 1

line_num_lang = {
    "1호선" : "Line 1",
    "2호선" : "Line 2",
    "3호선" : "Line 3",
    "4호선" : "Line 4",
    "5호선" : "Line 5",
    "6호선" : "Line 6",
    "7호선" : "Line 7",
    "8호선" : "Line 8",
}

log_data = []

for i in range(0,log_count,2) : #for문으로 0부터 metro_2019_log에 넣은 모든 열들의 갯수를 넣은 log_count의 값 -1 까지 
                                #2씩 증가하며 돌린다.  
    indata=metro_2019_log[i] #승차가 들어간 열을 indata에 차례로 넣는다.
    outdate=metro_2019_log[i+1] #하차가 들어간 열을 outdate에 차례로 넣는다.
    
    if(indata[0]==outdate[0] and indata[1]==outdate[1] and indata[2]==outdate[2]) :
        
        station_name = indata[3]
        dateTemp = indata[0]

        station_name_full = station_name

        # if '(' in station_name :
        if station_name.find("(") > 0 : #만약 station_name에 "("가 있다면 값이 1이 되어서 0보다 크게되고
            station_name=station_name.split("(")[0] #"("기준으로 자른 앞쪽 을 station_name에 저장한다 

        if(station_name == "총신대입구") : #총신대입구(이수)는 예외로 station_name을 이수로 바꾸어 준다.
          station_name = "이수"; 
          station_name_full = "총신대입구(이수)"

        for h in range(20) : #for문 돌려 4시부터 23시까지 dateTemp입력 된 날짜 데이터에 차례대로 추가해준다.
                             #원래는 5시 부터 24 혹은 0으로 까지를 넣고 싶으나 datetime은 23시까지만 출력가능...
                             #방법을 모르겠음.
            rdate=datetime.datetime(dateTemp.year,dateTemp.month,dateTemp.day,h+4)
            people_in = indata[h+5]
            people_out = outdate[h+5]

            # print(rdate)
            # print(people_in)
            # print(people_out)

            # print(station_name)

            log_raw = {}

            log_raw = {
                "@timestamp":rdate,
                "code":indata[2],
                "line_num":indata[1],
                "line_num_en" :line_num_lang[indata[1]],
                "station" : {
                    "name" : station_name_full,
                    "kr" : station_name
                },
                "location" : {
                    "lat" : subway_location[station_name][0],
                    "lon" : subway_location[station_name][1]
                },
                "people" : {
                    "in": people_in,
                    "out": people_out,
                    "totla": people_in+people_out
                }
            }

            log_data.append(log_raw)

class MyEncoder(json.JSONEncoder):
    def default(self, obj):
        return obj.isoformat()

with open('file.json', 'w', encoding="utf-8") as fp:
    fp.write(
        '[' +
        ',\n'.join(json.dumps(i,ensure_ascii=False,cls=MyEncoder) for i in log_data) + 
        #그리고 dumps()와 dump()는 파이썬 타입을 JSON으로 변환하는 메소드입니다.
        #파일 IO를 다루는 메소드는 s가 붙지 않고 문자열을 다루는 메소드는 s가 붙습니다.
        #파이썬의 json 모듈을 통해 별도의 추가 모듈 없이도 쉽고 편리하게 JSON을 다룰 수 있습니다.
        #encure_ascii=False를 해주어야 문자그대로를 출력할 수 있게 된다 만약 true이면 아스키코드가 출력된다.
        #log_data에 있는 값을 차례대로 인코딩을 거치어서 file.json에 변환하여 만든다.
        #위에라 [ 블라블라,
        #         블라블라,
        #       ]
        #이런식으로 저장할 수 있도록 앞뒤에 '[' ',\n, ']\n'을 넣어주었다.
        ']\n')'''